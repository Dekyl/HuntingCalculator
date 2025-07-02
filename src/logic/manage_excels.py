from datetime import datetime
import openpyxl, os, shutil
from openpyxl.utils import get_column_letter
from logic.logs import add_log

from config.config import saved_sessions_folder
from logic.data_classes.save_session_data import SaveSessionData

def delete_saved_session(file_path: str) -> int:
    """
    Delete a specific session file.
        :param file_path: The path to the session file to delete.
        :return: 0 if the deletion was successful, -1 if the file does not exist, -2 if an error occurred.
    """
    try:
        if not os.path.exists(file_path):
            add_log(f"Session file '{file_path}' does not exist.", "error")
            return -1
        os.remove(file_path)
        add_log(f"Session file '{file_path}' deleted successfully.", "info")
        return 0
    except Exception as e:
        add_log(f"Error deleting session file '{file_path}': {e}", "error")
        return -2

def clean_sessions() -> int:
    """
    Clean the sessions of the hunting session by all saved files.
    This function deletes the "Hunting Sessions" directory and all its contents,
    then recreates the directory to start fresh.
        :return: 1 if successful, 0 if folder not found or empty, -1 if folder not found, -2 if an unexpected error occurs.
    """
    try:
        if not os.path.exists(saved_sessions_folder):
            add_log(f"Folder {saved_sessions_folder} not found, creating it. No sessions were deleted", "info")
            os.mkdir(saved_sessions_folder)
            return -1
        elif os.path.isdir(saved_sessions_folder) and not os.listdir(saved_sessions_folder):
            add_log(f"Folder {saved_sessions_folder} is empty, nothing to delete", "info")
        elif os.path.isdir(saved_sessions_folder):
            add_log(f"Deleting all saved sessions in '{saved_sessions_folder}'", "info")
            shutil.rmtree(saved_sessions_folder)
            os.mkdir(saved_sessions_folder)
            add_log(f"Clean sessions dialog selection -> 1 (Success)", "info")
            return 1
        else:
            add_log(f"Unexpected error. '{saved_sessions_folder}' is not a directory or does not exist", "error")
            add_log(f"Clean sessions dialog selection -> -2 (Unexpected error)", "error")
            return -2

    except Exception as e:
        add_log(f"Unexpected error. Error while cleaning sessions: {e}", "error")
        add_log(f"Clean sessions dialog selection -> -2 (Unexpected error)", "error")
        return -2

    add_log(f"Clean sessions dialog selection -> 0 (No elements found to delete)", "info")
    return 0

class SaveSession:
    """ 
    Class to handle saving hunting session results to an Excel file.
    This class provides methods to save session data, including results and metadata,
    to an Excel file in a structured format.
    """
    def __init__(self, session_data: SaveSessionData):
        """
        Initialize the SaveSession instance with session data.
            :param session_data: An instance of SaveSessionData containing the results and metadata of the hunting session.
        """
        self.session_data = session_data
        self.name_spot = self.session_data.name_spot.lower().replace(" ", "_").replace("/", "_").replace("\\", "_")  # Sanitize the name_spot for file naming
        
        self.taxed_res_h = self.session_data.taxed_res_h
        self.taxed_res = self.session_data.taxed_res
        self.total_res = self.session_data.total_res
        self.total_res_h = self.session_data.total_res_h

        self.res_labels = self.session_data.res_name
        self.res_data = self.session_data.res_data
        self.hours = self.res_data[-1]  # Last element in res_data is the number of hours in str
        self.label_hours = self.res_labels[-1] # Last label in res_name is "Hours"
        self.hours_digit = 0  # Initialize hours_digit to 0, will be set later
        self.action_user = self.session_data.user_action  # User action string from the session data

    def save(self) -> bool:
        """
        Save the results of a hunting session to an Excel file.
            :return: False if an error occurs, otherwise True.
        """
        try:
            self.hours_digit = int(self.hours) # Ensure hours is an integer
        except:
            return False
        
        now = datetime.now()
        now_str = now.strftime("%d-%m-%Y_%H-%M-%S")
        

        results_hour = str(self.taxed_res_h / 1000000000)[:5]

        path = f"{saved_sessions_folder}/{now_str}_({results_hour}b).xlsx"

        workbook = openpyxl.Workbook()
        worksheet = workbook["Sheet"]
        worksheet.title = "Hunting Session Results"
        worksheet["A1"] = "Item Name"
        worksheet["B1"] = "Number of Items"

        max_label_width = len("Item Name")  # Initialize with the width of the header
        for i, label in enumerate(self.res_labels[:-1]):
            try:
                int(self.res_data[i])
            except:
                return False
            
            row = i + 2  # Start from row 2 for labels and data
            worksheet[f"A{row}"] = label
            cell_val = worksheet[f"B{row}"]
            cell_val.value = self.res_data[i]
            cell_val.number_format = '#,##0' # Format as number with thousands separator

            if len(label) > max_label_width:
                max_label_width = len(label)

        worksheet.column_dimensions['A'].width = max_label_width + 2

        start_column = 3
        worksheet[f"{get_column_letter(start_column)}1"] = self.label_hours  # Last label is "Hours"
        worksheet[f"{get_column_letter(start_column)}2"] = self.hours  # Last data input is the number of hours

        start_column += 1

        dic_results = [
            self.total_res,
            self.total_res_h,
            self.taxed_res,
            self.taxed_res_h
        ]

        for i, label in enumerate(self.session_data.labels_res):
            cell_val = worksheet[get_column_letter(start_column) + "2"]
            worksheet[get_column_letter(start_column) + "1"] = label
            cell_val.value = dic_results[i]
            cell_val.number_format = '#,##0' # Format as number with thousands separator
            start_column += 1

        worksheet[f"{get_column_letter(start_column)}1"] = "User action"  # Last label is "Hours"
        worksheet[f"{get_column_letter(start_column)}2"] = self.action_user  # Last data input is the number of hours

        workbook.save(filename=path)

        log_message = (
            f"\tResults saved in '{path}'\n"
            f"\tInput Labels: {self.res_labels}\n"
            f"\tInput Data: {self.res_data}\n"
            f"\tResult Labels: {self.session_data.labels_res}\n"
            f"\tTotal Results: {self.total_res:,}\n"
            f"\tResults per Hour: {self.total_res_h:,}\n"
            f"\tResults after Tax: {self.taxed_res:,}\n"
            f"\tResults after Tax per Hour: {self.taxed_res_h:,}\n"
            f"\tAction user: {self.action_user}"
        )
        add_log(log_message, "info")

        return self.save_average()  # Call the save_average method to update the average results

    def save_average(self) -> bool:
        """
        Updates the average results of hunting sessions of a spot saved into the excel file.
            :return: True if successful, False if an error occurs.
        """
        path = f"{saved_sessions_folder}/average_results_{self.name_spot}.xlsx"

        if not os.path.exists(path): # If the file does not exist, create it
            if self.hours_digit <= 0 or len(self.res_labels) != len(self.res_data):
                return False
            workbook = openpyxl.Workbook()
            worksheet = workbook["Sheet"]

            worksheet["A1"] = "Item Name"
            worksheet["B1"] = "Total Number of Items"
            worksheet["C1"] = "Average Number of Items"
            worksheet["D1"] = "Total Hours"
            worksheet["D2"] = self.hours_digit  # Number of hours
            worksheet["E1"] = "Total Sessions"
            worksheet["E2"] = 1  # Number of sessions
            worksheet["F1"] = "Average Hours/Session"
            worksheet["F2"] = self.hours_digit  # Number of hours
            worksheet["G1"] = "Total Money" # Total money acquired and taxed
            worksheet["G2"] = self.taxed_res_h * self.hours_digit
            worksheet["H1"] = "Average Money/Hour"
            worksheet["H2"] = self.taxed_res_h
            worksheet["I1"] = "Average Money/Session"
            worksheet["I2"] = self.taxed_res_h * self.hours_digit / self.hours_digit

            max_label_width = len("Item Name")  # Initialize with the width of the header
            for i, label in enumerate(self.res_labels[:-1]):
                row = i + 2  # Start from row 2 for labels and data
                worksheet[f"A{row}"] = label[0:label.rfind(" ")]
                worksheet[f"B{row}"] = self.res_data[i]
                worksheet[f"C{row}"] = self.res_data[i]

                if len(label) > max_label_width:
                    max_label_width = len(label)

            worksheet.column_dimensions['A'].width = max_label_width + 2

            workbook.save(filename=path)
            return True
        
        workbook = openpyxl.load_workbook(filename=path)
        worksheet = workbook["Sheet"]

        total_hours = int(worksheet['D2'].value) # Total hours from the previous sessions
        if total_hours <= 0:
            return False

        new_total_money = int(worksheet['G2'].value) + self.taxed_res_h * self.hours_digit
        new_total_hours = total_hours + self.hours_digit
        new_total_sessions = int(worksheet['E2'].value) + 1

        worksheet['D2'] = new_total_hours
        worksheet['E2'] = new_total_sessions
        worksheet['F2'] = float(new_total_hours / new_total_sessions)
        worksheet['F2'].number_format = '#,##0.00' # Format as number with two decimal places
        worksheet['G2'] = new_total_money
        worksheet['H2'] = int(new_total_money / new_total_hours)
        worksheet['I2'] = int(new_total_money / new_total_sessions)

        for i, data_inp in enumerate(self.res_data[:-1]):
            row = i + 2  # Start from row 2 for labels and data

            cell_val_total = worksheet[f"B{row}"]
            cell_val_total.value = int(cell_val_total.value) + int(data_inp) if cell_val_total.value else int(data_inp)
            cell_val_total.number_format = '#,##0' # Format as number with thousands separator

            cell_val_average = worksheet[f"C{row}"]
            cell_val_average.value = float(cell_val_total.value / new_total_hours) if new_total_hours > 0 else 0
            cell_val_average.number_format = '#,##0.00' # Format as number with thousands separator

        workbook.save(filename=path)
        log_message = (
            f"\tAverage results updated for spot: {self.name_spot}\n"
            f"\tResults saved in {path}\n"
            f"\tInput Labels: {self.res_labels}\n"
            f"\tInput Data: {self.res_data}\n"
            f"\tTotal Hours: {self.hours_digit}\n"
            f"\tTotal Money: {self.taxed_res_h * self.hours_digit:,}\n"
            f"\tAverage Money/Hour: {self.taxed_res_h:,}\n"
            f"\tAverage Money/Session: {self.taxed_res_h * self.hours_digit / self.hours_digit:,}"
        )
        add_log(log_message, "info")
        return True