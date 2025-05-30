from datetime import datetime
import openpyxl, os

def save_excel(labels_input, data_input, labels_res, results_tot, results_tot_h, results_tax, results_tax_h):
    try:
        int(data_input[-1])
    except:
        return -1
    
    now = datetime.now()
    now_str = now.strftime("%d-%m-%Y_%H-%M-%S")

    results_hour = str(results_tax_h / 1000000000)[:5]

    path = './Hunting Sessions/' + now_str + "_(" + results_hour + "b).xlsx"

    workbook = openpyxl.Workbook()
    worksheet = workbook["Sheet"]

    for i, label in enumerate(labels_input):
        cell_labels = chr(65 + i) + "1"
        cell_data = chr(65 + i) + "2"

        worksheet[cell_labels] = label

        try:
            int(data_input[i])
        except:
            return -1
        
        worksheet[cell_data] = data_input[i]

    dic_results = {
        0: results_tot,
        1: results_tot_h,
        2: results_tax,
        3: results_tax_h
    }

    for i, label in enumerate(labels_res):
        cell_labels_res = chr(65 + i) + "4"
        cell_data_res = chr(65 + i) + "5"

        worksheet[cell_labels_res] = label
        worksheet[cell_data_res] = dic_results[i]

    workbook.save(filename=path)
    return save_average(results_tax_h, int(data_input[-1]), labels_input, data_input)

def save_average(results_tax_h, hours_session, labels_input, data_input):
    path = "./Hunting Sessions/average_results.xlsx"

    if not os.path.exists(path):
        workbook = openpyxl.Workbook()
        worksheet = workbook["Sheet"]

        worksheet["A1"] = "Total Hours"
        worksheet["B1"] = "Total Sessions"
        worksheet["C1"] = "Average Money/hour"
        worksheet["D1"] = "Average Hours/session"
        worksheet["E1"] = "Total Money"

        worksheet["A2"] = hours_session
        worksheet["B2"] = 1
        worksheet["C2"] = results_tax_h
        worksheet["D2"] = hours_session
        worksheet["E2"] = results_tax_h*hours_session

        worksheet['A5'] = "Total"
        worksheet['A7'] = "Average/hour"

        for i, label in enumerate(labels_input):
            cell_labels = chr(65 + i) + "4"
            cell_data = chr(65 + i) + "6"
            cell_data_average = chr(65 + i) + "8"

            worksheet[cell_labels] = label[0:label.rfind(" ")]

            try:
                int(data_input[i])
            except:
                return -1

            worksheet[cell_data] = int(data_input[i])

            if (i < len(labels_input)-1):
                worksheet[cell_data_average] = int(data_input[i]) / hours_session

        workbook.save(filename=path)
        return
    
    workbook = openpyxl.load_workbook(filename=path)
    worksheet = workbook["Sheet"]

    total_hours = int(worksheet['A2'].value)

    if total_hours <= 0:
        return -1

    new_total_money = int(worksheet['E2'].value) + results_tax_h*hours_session
    new_total_hours = total_hours + hours_session
    new_total_sessions = int(worksheet['B2'].value) + 1

    worksheet['A2'] = new_total_hours
    worksheet['B2'] = new_total_sessions
    worksheet['C2'] = int(new_total_money / new_total_hours)
    worksheet['D2'] = new_total_hours / new_total_sessions
    worksheet['E2'] = new_total_money

    worksheet["A1"] = "Total Hours"
    worksheet["B1"] = "Total Sessions"
    worksheet["C1"] = "Average Money/hour"
    worksheet["D1"] = "Average Hours/session"
    worksheet["E1"] = "Total Money"

    worksheet['A5'] = "Total"
    worksheet['A7'] = "Average/hour"

    for i, label in enumerate(labels_input):
        cell_labels = chr(65 + i) + "4"
        cell_data = chr(65 + i) + "6"
        cell_data_average = chr(65 + i) + "8"

        worksheet[cell_labels] = label[0:label.rfind(" ")]

        try:
            int(data_input[i])
            int(worksheet[cell_data].value)
        except:
            return -1
        
        actual_item_total = int(worksheet[cell_data].value)
        new_item_total = actual_item_total + int(data_input[i])

        worksheet[cell_data] = int(new_item_total)
        if (i < len(labels_input)-1):
                worksheet[cell_data_average] = new_item_total / new_total_hours

    workbook.save(filename=path)
